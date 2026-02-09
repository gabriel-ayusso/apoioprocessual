import io
from uuid import UUID
from datetime import date, datetime
from typing import Optional, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.models import Transacao, Evento, Document, Processo


class ExcelGenerator:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def generate(
        self,
        processo_id: UUID,
        tipo: str,
        data_inicio: Optional[date] = None,
        data_fim: Optional[date] = None,
        categorias: Optional[List[str]] = None,
        pagadores: Optional[List[str]] = None,
    ) -> tuple[str, bytes]:
        """Generate Excel report and return filename and content."""

        if tipo == "transacoes":
            return await self._generate_transacoes(
                processo_id, data_inicio, data_fim, categorias, pagadores
            )
        elif tipo == "timeline":
            return await self._generate_timeline(processo_id, data_inicio, data_fim)
        elif tipo == "evidencias":
            return await self._generate_evidencias(processo_id, categorias)
        else:
            raise ValueError(f"Tipo de relatorio invalido: {tipo}")

    def _create_workbook(self, title: str) -> Workbook:
        """Create a styled workbook."""
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel limit

        return wb

    def _style_header(self, ws, row: int, columns: List[str]):
        """Style header row."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, value in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _auto_width(self, ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    async def _generate_transacoes(
        self,
        processo_id: UUID,
        data_inicio: Optional[date],
        data_fim: Optional[date],
        categorias: Optional[List[str]],
        pagadores: Optional[List[str]],
    ) -> tuple[str, bytes]:
        """Generate transactions report."""
        # Get processo info
        processo_result = await self.db.execute(
            select(Processo).where(Processo.id == processo_id)
        )
        processo = processo_result.scalar_one()

        # Build query
        query = select(Transacao).where(Transacao.processo_id == processo_id)

        if data_inicio:
            query = query.where(Transacao.data >= data_inicio)
        if data_fim:
            query = query.where(Transacao.data <= data_fim)
        if categorias:
            query = query.where(Transacao.categoria.in_(categorias))
        if pagadores:
            query = query.where(Transacao.pagador.in_(pagadores))

        query = query.order_by(Transacao.data.desc().nullslast())

        result = await self.db.execute(query)
        transacoes = result.scalars().all()

        # Create workbook
        wb = self._create_workbook("Transacoes")
        ws = wb.active

        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = f"Relatorio de Transacoes - {processo.titulo}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Metadata
        ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A3"] = f"Periodo: {data_inicio or 'Inicio'} a {data_fim or 'Fim'}"

        # Headers
        columns = ["Data", "Descricao", "Valor", "Pagador", "Beneficiario", "Categoria", "Confianca", "Revisado"]
        self._style_header(ws, 5, columns)

        # Data
        for row, trans in enumerate(transacoes, 6):
            ws.cell(row=row, column=1, value=str(trans.data) if trans.data else "")
            ws.cell(row=row, column=2, value=trans.descricao)
            ws.cell(row=row, column=3, value=float(trans.valor) if trans.valor else 0)
            ws.cell(row=row, column=4, value=trans.pagador or "")
            ws.cell(row=row, column=5, value=trans.beneficiario or "")
            ws.cell(row=row, column=6, value=trans.categoria or "")
            ws.cell(row=row, column=7, value=f"{trans.confianca:.0%}" if trans.confianca else "")
            ws.cell(row=row, column=8, value="Sim" if trans.revisado_humano else "Nao")

        # Summary row
        summary_row = len(transacoes) + 7
        ws.cell(row=summary_row, column=1, value="TOTAL")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        ws.cell(row=summary_row, column=3, value=sum(float(t.valor or 0) for t in transacoes))
        ws.cell(row=summary_row, column=3).font = Font(bold=True)

        self._auto_width(ws)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"transacoes_{processo_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return filename, output.getvalue()

    async def _generate_timeline(
        self,
        processo_id: UUID,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> tuple[str, bytes]:
        """Generate timeline report."""
        # Get processo info
        processo_result = await self.db.execute(
            select(Processo).where(Processo.id == processo_id)
        )
        processo = processo_result.scalar_one()

        # Build query
        query = select(Evento).where(Evento.processo_id == processo_id)

        if data_inicio:
            query = query.where(Evento.data >= data_inicio)
        if data_fim:
            query = query.where(Evento.data <= data_fim)

        query = query.order_by(Evento.data)

        result = await self.db.execute(query)
        eventos = result.scalars().all()

        # Create workbook
        wb = self._create_workbook("Timeline")
        ws = wb.active

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = f"Timeline de Eventos - {processo.titulo}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Headers
        columns = ["Data", "Tipo", "Descricao", "Importancia", "Evidencia", "Confianca"]
        self._style_header(ws, 3, columns)

        # Data
        for row, evento in enumerate(eventos, 4):
            ws.cell(row=row, column=1, value=str(evento.data))
            ws.cell(row=row, column=2, value=evento.tipo or "")
            ws.cell(row=row, column=3, value=evento.descricao)
            ws.cell(row=row, column=4, value=evento.importancia)
            ws.cell(row=row, column=5, value=(evento.trecho_evidencia or "")[:100])
            ws.cell(row=row, column=6, value=f"{evento.confianca:.0%}" if evento.confianca else "")

        self._auto_width(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"timeline_{processo_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return filename, output.getvalue()

    async def _generate_evidencias(
        self,
        processo_id: UUID,
        categorias: Optional[List[str]],
    ) -> tuple[str, bytes]:
        """Generate evidence report."""
        # Get processo info
        processo_result = await self.db.execute(
            select(Processo).where(Processo.id == processo_id)
        )
        processo = processo_result.scalar_one()

        # Get documents
        query = select(Document).where(
            Document.processo_id == processo_id,
            Document.status == "processed",
        )

        if categorias:
            query = query.where(Document.tipo.in_(categorias))

        query = query.order_by(Document.data_referencia.desc().nullslast())

        result = await self.db.execute(query)
        documents = result.scalars().all()

        # Create workbook
        wb = self._create_workbook("Evidencias")
        ws = wb.active

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = f"Relatorio de Evidencias - {processo.titulo}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Headers
        columns = ["Data Ref", "Tipo", "Titulo", "Descricao", "Participantes", "Arquivo", "Texto Extraido (resumo)"]
        self._style_header(ws, 3, columns)

        # Data
        for row, doc in enumerate(documents, 4):
            ws.cell(row=row, column=1, value=str(doc.data_referencia) if doc.data_referencia else "")
            ws.cell(row=row, column=2, value=doc.tipo)
            ws.cell(row=row, column=3, value=doc.titulo)
            ws.cell(row=row, column=4, value=doc.descricao or "")
            ws.cell(row=row, column=5, value=", ".join(doc.participantes) if doc.participantes else "")
            ws.cell(row=row, column=6, value=doc.arquivo_nome)
            ws.cell(row=row, column=7, value=(doc.texto_extraido or "")[:200] + "..." if doc.texto_extraido and len(doc.texto_extraido) > 200 else doc.texto_extraido or "")

        self._auto_width(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"evidencias_{processo_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return filename, output.getvalue()
